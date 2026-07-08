"""
pds_generator.py - Génération du livrable PDS (plan de soudure Excel).

Produit un classeur .xlsx à partir d'un GABARIT (PDS_template.xlsx) : un onglet
par boîte BPE, nommé comme la BPE, rempli depuis les couches BPE, CABLES, BTS, PT.

PRINCIPE D'INDÉPENDANCE : ce module ne connaît ni l'interface web ni la base CRM.
Il reçoit des GeoDataFrames + un chemin de gabarit et retourne un chemin de fichier.
"""

import os
import math
import logging

import openpyxl
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger("crm_sig.pds")

# Code couleur standard des fibres/tubes (position 1..12)
COULEURS = ["rouge", "bleu", "vert", "jaune", "violet", "blanc",
            "orange", "gris", "marron", "noir", "turquoise", "rose"]

# Fonds de couleur (ARGB) repris À L'IDENTIQUE du fichier de référence.
COULEUR_FILL = {
    "rouge": "FFFF0000", "bleu": "FF0000FF", "vert": "FF00FF00", "jaune": "FFFFFF00",
    "violet": "FF6633FF", "blanc": "FFFFFFFF", "orange": "FFFF8C00", "gris": "FF808080",
    "marron": "FF8B4513", "noir": "FF000000", "turquoise": "FF40E0D0", "rose": "FFFFC0CB",
}
FILL_ETAT_E = "FFFFC000"     # état "E" (épissure) sur fond ambre ; "ST" reste sans fond
TAB_EN_SERVICE = "FFFFC000"  # couleur d'onglet BPE EN SERVICE
TAB_DEFAUT = "FFFF0000"      # couleur d'onglet par défaut (EN ETUDE)


def _fill(argb):
    """PatternFill solide de la couleur donnée (ou aucun fond si argb est None)."""
    return PatternFill(fill_type="solid", fgColor=argb) if argb else PatternFill(fill_type=None)


# Modèles de BPE valides (= liste déroulante J4:K4). Certains sont "autonomes"
# (une seule référence suffit, ex. TENIOPEOC8144FR5), d'autres = MODELE + REFERENCE.
MODELES_BPE = [
    "3M T0 MFO13280-1",
    "3M T1 N501733A",
    "OFDC-B8-S36-2-NN8",
    "TENIOPEOC8144FR5",
]


def _modele_bpe(modele, reference):
    """Valeur du champ 'Modèle BPE' (J4) : si la référence (ou le modèle) est un code
    autonome connu, on l'utilise seul ; sinon MODELE + ' ' + REFERENCE."""
    modele = (modele or "").strip()
    reference = (reference or "").strip()
    combine = (modele + " " + reference).strip()
    connus = {m.upper(): m for m in MODELES_BPE}
    for cand in (combine, reference, modele):
        if cand and cand.upper() in connus:
            return connus[cand.upper()]
    return combine

# Disposition d'une cassette dans le gabarit : titre en 7+(c-1)*15, 12 fibres en 9..20
CASSETTE_PAS = 15
PREMIERE_FIBRE = 9

CARACTERES_INTERDITS = set(r':\/?*[]')


def _nom_onglet(nom: str) -> str:
    """Nettoie un nom pour en faire un nom d'onglet Excel valide (<= 31 car.)."""
    propre = "".join(c for c in str(nom) if c not in CARACTERES_INTERDITS).strip()
    return (propre or "BPE")[:31]


def _harmoniser(gdf, crs):
    if gdf is None or len(gdf) == 0 or crs is None:
        return gdf
    if gdf.crs is not None and gdf.crs != crs:
        return gdf.to_crs(crs)
    return gdf


def _point(geom):
    """Retourne un point représentatif d'une géométrie (point ou centroïde)."""
    if geom is None or geom.is_empty:
        return None
    return geom if geom.geom_type == "Point" else geom.centroid


def _coords_ligne(g):
    """Liste plate des coordonnées d'une ligne — gère LineString ET
    MultiLineString (``.coords`` échoue sur les géométries multi-parties :
    « Sub-geometries may have coordinate sequences, but multi-part geometries
    do not »). c[0] / c[-1] restent les extrémités globales de la ligne."""
    if g is None:
        return []
    if getattr(g, "geom_type", "") == "MultiLineString":
        out = []
        for part in g.geoms:
            out.extend(part.coords)
        return out
    return list(g.coords)


def _capacite_map(cables_gdf, defaut):
    """Dictionnaire nom_cable -> capacité (FO), avec repli sur `defaut`."""
    m = {}
    if cables_gdf is None:
        return m
    for _, r in cables_gdf.iterrows():
        nom = r.get("NOM")
        capa = r.get("CAPACITE")
        try:
            capa = int(float(capa))
            if capa <= 0:
                capa = defaut
        except (TypeError, ValueError):
            capa = defaut
        if nom:
            m[str(nom)] = capa
    return m


def _ordonner_boites(bpe_gdf, cables_gdf, bts_gdf, tol=3.0, plafond=100.0):
    """
    Pour chaque BPE, détermine le câble ENTRANT et SORTANT à partir des câbles
    INCIDENTS, quel que soit le nombre de boîtes.

    Rattachement des câbles à une BPE :
      1) câbles dont une extrémité est <= tol (3 m) de la BPE (boîte sur le réseau) ;
      2) sinon repli : câbles du NŒUD-CÂBLE le plus proche (jonction d'extrémités),
         dans un plafond (défaut 100 m) — gère les BPE décalées du réseau (ex. FM029
         posée à ~30 m du poteau où les câbles se rejoignent) ; au-delà -> aucun.

    Entrant/sortant :
      - 0 câble  : entrant = sortant = None
      - 1 câble  : entrant = sortant = ce câble (boîte terminale / origine)
      - 2+ câbles: sortant = câble le plus proche du BTS (sens vers le site),
        entrant = le plus éloigné. Sans BTS : l'adduction (préfixe CAD) = sortant.

    Retourne une liste de dict : {bpe (row), entrant, sortant}.
    """
    from shapely.geometry import Point

    bts_pt = None
    if bts_gdf is not None and len(bts_gdf) > 0:
        bts_pt = _point(bts_gdf.iloc[0].geometry)

    cables = []
    if cables_gdf is not None:
        for _, rc in cables_gdf.iterrows():
            g = rc.geometry
            if g is not None and not g.is_empty:
                cables.append((str(rc.get("NOM")), g))

    # Nœuds du réseau câble : extrémités regroupées (tolérance 1 m) -> {nom: geom}
    noeuds = []  # [ [Point, {nom: geom}] ]

    def _ajouter_extremite(pt, nom, geom):
        for nd in noeuds:
            if nd[0].distance(pt) <= 1.0:
                nd[1].setdefault(nom, geom)
                return
        noeuds.append([pt, {nom: geom}])

    for nom, g in cables:
        c = _coords_ligne(g)
        _ajouter_extremite(Point(c[0][:2]), nom, g)
        _ajouter_extremite(Point(c[-1][:2]), nom, g)

    resultats = []
    for _, rb in bpe_gdf.iterrows():
        bp = _point(rb.geometry)
        if bp is None:
            resultats.append({"bpe": rb, "entrant": None, "sortant": None})
            continue

        # 1) câbles incidents (extrémité <= tol de la BPE)
        incident = {}
        for nom, g in cables:
            c = _coords_ligne(g)
            if min(Point(c[0][:2]).distance(bp), Point(c[-1][:2]).distance(bp)) <= tol:
                incident.setdefault(nom, g)

        # 2) repli : nœud-câble le plus proche dans le plafond
        if not incident and noeuds:
            nd = min(noeuds, key=lambda n: n[0].distance(bp))
            d = nd[0].distance(bp)
            if d <= plafond:
                incident = dict(nd[1])
                logger.info(f"BPE '{rb.get('NOM')}' rattachée au nœud câble le plus proche ({d:.1f} m).")

        noms = list(incident.keys())
        if not noms:
            entrant = sortant = None
        elif len(noms) == 1:
            entrant = sortant = noms[0]
        else:
            if bts_pt is not None:
                tri = sorted(noms, key=lambda n: incident[n].distance(bts_pt))
            else:
                # sans BTS : l'adduction (CAD) part vers le site -> sortant
                tri = sorted(noms, key=lambda n: (not n.upper().startswith("CAD"), n))
            sortant, entrant = tri[0], tri[-1]

        resultats.append({"bpe": rb, "entrant": entrant, "sortant": sortant})
    return resultats


def _remplir_onglet(ws, box, chamber, capa_entrant, sit_fm, nb_fo, commentaire):
    """Remplit un onglet (copie du gabarit) pour une boîte."""
    bpe = box["bpe"]
    adresse = str(bpe.get("ADRESSE") or "").strip()
    cp = str(bpe.get("CP") or "").strip()
    ville = str(bpe.get("VILLE") or "").strip()
    modele = str(bpe.get("MODELE") or "").strip()
    reference = str(bpe.get("REFERENCE") or "").strip()

    ws["B2"] = chamber or ""
    ws["B3"] = f"{adresse}\n{cp} {ville}".strip()
    _al = ws["B3"].alignment
    ws["B3"].alignment = Alignment(horizontal=_al.horizontal, vertical=_al.vertical or "center", wrap_text=True)
    ws["B4"] = str(bpe.get("NOM") or "")
    ws["F2"] = commentaire or f"Pose d'une BPE et soudures de {nb_fo} FO"
    ws["F3"] = "NON"
    ws["G4"] = capa_entrant
    ws["J4"] = _modele_bpe(modele, reference)

    # Couleur d'onglet reprise de la référence (EN SERVICE = ambre, sinon rouge)
    etat_bpe = str(bpe.get("ETAT") or "").upper()
    ws.sheet_properties.tabColor = TAB_EN_SERVICE if "SERVICE" in etat_bpe else TAB_DEFAUT

    # Listes déroulantes (non copiées par copy_worksheet) : F3 = OUI/NON, J4:K4 = modèle BPE
    dv = DataValidation(type="list", formula1='"OUI,NON"', allow_blank=True)
    ws.add_data_validation(dv); dv.add(ws["F3"])
    dv_mod = DataValidation(
        type="list", allow_blank=True,
        formula1='"' + ",".join(MODELES_BPE) + '"')
    ws.add_data_validation(dv_mod); dv_mod.add("J4:K4")

    # Corps : plan de soudure
    nb_tubes = max(1, math.ceil(capa_entrant / 12))
    entrant = box["entrant"] or ""
    sortant = box["sortant"] or entrant
    for t in range(1, nb_tubes + 1):
        base = PREMIERE_FIBRE + (t - 1) * CASSETTE_PAS
        tube_color = COULEURS[(t - 1) % 12]
        for p in range(1, 13):
            r = base + (p - 1)
            g = (t - 1) * 12 + p  # index global de la fibre
            fibre_color = COULEURS[(p - 1) % 12]
            ws[f"B{r}"] = entrant
            ws[f"C{r}"] = t
            ws[f"D{r}"] = tube_color; ws[f"D{r}"].fill = _fill(COULEUR_FILL.get(tube_color))
            ws[f"E{r}"] = 1
            ws[f"F{r}"] = p
            ws[f"G{r}"] = fibre_color; ws[f"G{r}"].fill = _fill(COULEUR_FILL.get(fibre_color))
            if g <= nb_fo:
                ws[f"H{r}"] = "E"; ws[f"H{r}"].fill = _fill(FILL_ETAT_E)
                ws[f"I{r}"] = fibre_color; ws[f"I{r}"].fill = _fill(COULEUR_FILL.get(fibre_color))
                ws[f"J{r}"] = p
                ws[f"K{r}"] = 1
                ws[f"L{r}"] = tube_color; ws[f"L{r}"].fill = _fill(COULEUR_FILL.get(tube_color))
                ws[f"M{r}"] = t
                ws[f"N{r}"] = sortant
                if g == 1:
                    ws[f"O{r}"] = sit_fm
            else:
                ws[f"H{r}"] = "ST"


def generer_pds(bpe_gdf, cables_gdf, bts_gdf, pt_gdf, chemin_sortie,
                template_path, nb_fo=2, capacite_defaut=48, commentaire=None):
    """
    Génère le classeur PDS et l'écrit dans `chemin_sortie`. Retourne (chemin, nb_onglets).
    """
    if bpe_gdf is None or len(bpe_gdf) == 0:
        raise ValueError("Couche BPE vide : impossible de générer le PDS.")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Gabarit PDS introuvable : {template_path}")

    crs = bpe_gdf.crs
    cables_gdf = _harmoniser(cables_gdf, crs)
    bts_gdf = _harmoniser(bts_gdf, crs)
    pt_gdf = _harmoniser(pt_gdf, crs)

    capa_map = _capacite_map(cables_gdf, capacite_defaut)
    sit_fm = ""
    if bts_gdf is not None and len(bts_gdf) > 0:
        sit_fm = str(bts_gdf.iloc[0].get("REF_PHFM") or "").strip()

    boites = _ordonner_boites(bpe_gdf, cables_gdf, bts_gdf)

    def chambre_de(bpe_row):
        """Nom du PT le plus proche de la BPE."""
        if pt_gdf is None or len(pt_gdf) == 0:
            return ""
        p = _point(bpe_row.geometry)
        if p is None:
            return ""
        best, bestd = "", None
        for _, rpt in pt_gdf.iterrows():
            nom = str(rpt.get("NOM") or "").strip()
            if not nom:  # ignorer les PT sans nom (coffrets, etc.)
                continue
            gp = _point(rpt.geometry)
            if gp is None:
                continue
            d = p.distance(gp)
            if bestd is None or d < bestd:
                bestd, best = d, nom
        return best

    wb = openpyxl.load_workbook(template_path)
    if "PDS" not in wb.sheetnames:
        raise ValueError("Le gabarit PDS doit contenir un onglet 'PDS'.")
    gabarit = wb["PDS"]

    # On part du fichier de référence tel quel : on retire les onglets-boîtes
    # d'exemple qu'il pourrait contenir (onglets VISIBLES autres que le gabarit),
    # tout en conservant le gabarit 'PDS' et les onglets techniques masqués.
    for ws in list(wb.worksheets):
        if ws.title != "PDS" and ws.sheet_state == "visible":
            del wb[ws.title]

    noms_utilises = set()
    nb = 0
    for box in boites:
        bpe = box["bpe"]
        nom_bpe = str(bpe.get("NOM") or f"BPE_{nb+1}")
        nom = _nom_onglet(nom_bpe)
        base_nom = nom; k = 1
        while nom in noms_utilises:
            k += 1; nom = f"{base_nom[:28]}_{k}"
        noms_utilises.add(nom)

        ws = wb.copy_worksheet(gabarit)
        ws.title = nom
        ws.sheet_state = "visible"

        capa_entrant = capa_map.get(box["entrant"], capacite_defaut)
        _remplir_onglet(ws, box, chambre_de(bpe), capa_entrant, sit_fm, nb_fo, commentaire)
        nb += 1

    # Le premier onglet-boîte devient l'onglet actif à l'ouverture
    for i, ws in enumerate(wb.worksheets):
        if ws.sheet_state == "visible" and ws.title != "PDS":
            wb.active = i
            break

    os.makedirs(os.path.dirname(chemin_sortie), exist_ok=True)
    wb.save(chemin_sortie)

    # Ré-injecter les boutons (contrôles de formulaire) supprimés par openpyxl
    try:
        from app.reporting import pds_controls
        pds_controls.injecter_controles(chemin_sortie, template_path)
    except Exception as e:
        logger.warning(f"Greffe des boutons échouée (PDS généré sans boutons) : {e}")

    logger.info(f"PDS généré : {chemin_sortie} ({nb} onglets)")
    return chemin_sortie, nb
