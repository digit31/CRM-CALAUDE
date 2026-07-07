"""
annexe_appuis.py - Annexes C6 / C7 (appuis Orange) et jointure à la couche PT.

Rôle
----
Les annexes fournissent la NATURE DES TRAVAUX par appui (Remplacement / Recalage /
Renforcement / Irremplaçable), non déductible du SHP :

  · C6 (feuille « Export 1 ») = source de vérité par appui : N° appui + Nature des
    travaux + type avant/après + nombre de poteau commandé + Lat/Lon (WGS84, DMS).
  · C7 (feuille « Commande »)  = récapitulatif des poteaux commandés (contrôle /
    complément des remplacements).

Principes
---------
  · Jointure PT ↔ appui : par NUMÉRO (``PT.NOM``) d'abord, sinon par GÉOLOCALISATION
    (Lat/Lon C6 → plus proche poteau PT dans une tolérance).
  · Enrichissement du PT livrable : on ne met à jour QUE ``NOM`` + ``CODE`` — la
    structure (schéma) du SHP est strictement préservée.
  · La nature des travaux N'EST PAS écrite dans le PT : elle est fournie à la volée
    (``natures_par_nom``) pour la Console, le rapport PDF et la symbologie.
"""

import re
import logging

logger = logging.getLogger("crm_sig.annexe_appuis")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _num(v):
    """Numéro d'appui normalisé (sans espaces ; « 50895.0 » -> « 50895 »)."""
    if v is None:
        return ""
    s = re.sub(r"\s+", "", str(v))
    m = re.fullmatch(r"(\d+)\.0+", s)   # flottant entier (lecture DBF numérique)
    return m.group(1) if m else s


def _nums_candidats(row):
    """Numéros candidats d'un poteau pour la jointure : NOM, CODE, et leur dernier
    groupe de chiffres (couvre « FT_68218_50895 » -> « 50895 »)."""
    cands = []
    for champ in ("NOM", "CODE"):
        v = _num(row.get(champ))
        if not v:
            continue
        if v not in cands:
            cands.append(v)
        grp = re.findall(r"\d+", v)
        if grp and grp[-1] not in cands:
            cands.append(grp[-1])
    return cands


def _norm_nature(v):
    """Normalise la nature des travaux vers un jeu de valeurs stable."""
    s = (str(v) if v is not None else "").strip().upper()
    if not s:
        return ""
    if "REMPLAC" in s:
        return "REMPLACEMENT"
    if "RECAL" in s:
        return "RECALAGE"
    if "RENFORC" in s:
        return "RENFORCEMENT"
    if "IRREMPLAC" in s:
        return "IRREMPLACABLE"
    return s


def _parse_dms(s):
    """« N 47°44'13,5" » / « E 007°16'05,5" » -> degrés décimaux (float) ou None."""
    if s is None:
        return None
    txt = str(s).strip().replace(",", ".")
    m = re.search(r"([NSEWO])?\s*(\d+)\s*[°:\s]\s*(\d+)\s*['′:\s]\s*([\d.]+)\s*[\"″]?\s*([NSEWO])?",
                  txt, re.IGNORECASE)
    if not m:
        # tentative décimale simple
        try:
            return float(txt)
        except ValueError:
            return None
    hemi = (m.group(1) or m.group(5) or "N").upper()
    deg = float(m.group(2)) + float(m.group(3)) / 60.0 + float(m.group(4)) / 3600.0
    if hemi in ("S", "W", "O"):
        deg = -deg
    return deg


def _index_entete(rows, max_scan=40):
    """Indice (0-based) de la ligne d'en-tête du tableau d'appuis (col « N° appui »)."""
    for i, r in enumerate(rows[:max_scan]):
        for c in (r or [])[:3]:
            t = str(c or "").strip().lower()
            if "appui" in t and (t.startswith("n") or "°" in t):
                return i
    return None


# ---------------------------------------------------------------------------
# Lecture des annexes
# ---------------------------------------------------------------------------

def lire_c6(chemin):
    """C6 « Export 1 » -> {num_appui: {nature, avant, apres, nb, lat, lon}}."""
    import openpyxl
    out = {}
    try:
        wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
    except Exception as e:
        logger.warning(f"C6 illisible ({chemin}) : {e}")
        return out
    try:
        ws = wb["Export 1"] if "Export 1" in wb.sheetnames else wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        h = _index_entete(rows)
        h = 7 if h is None else h
        for r in rows[h + 1:]:
            if not r:
                continue
            num = _num(r[0]) if len(r) > 0 else ""
            if not num or num.lower().startswith("n"):
                continue
            def g(i):
                return r[i] if len(r) > i else None
            out[num] = {
                "nature": _norm_nature(g(31)),
                "avant": str(g(30) or "").strip(),
                "apres": str(g(32) or "").strip(),
                "nb": g(33),
                "lat": _parse_dms(g(3)),
                "lon": _parse_dms(g(4)),
                "source": "C6",
            }
    finally:
        wb.close()
    return out


def lire_c7(chemin):
    """C7 « Commande » -> {num_appui: {nature, avant, apres}}."""
    import openpyxl
    out = {}
    try:
        wb = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
    except Exception as e:
        logger.warning(f"C7 illisible ({chemin}) : {e}")
        return out
    try:
        ws = wb["Commande"] if "Commande" in wb.sheetnames else wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        h = _index_entete(rows)
        h = 16 if h is None else h
        for r in rows[h + 1:]:
            if not r:
                continue
            num = _num(r[0]) if len(r) > 0 else ""
            if not num or num.lower().startswith("n"):
                continue
            def g(i):
                return r[i] if len(r) > i else None
            out[num] = {
                "nature": _norm_nature(g(2)),
                "avant": str(g(1) or "").strip(),
                "apres": str(g(3) or "").strip(),
                "source": "C7",
            }
    finally:
        wb.close()
    return out


def charger_annexes(chemin_c6=None, chemin_c7=None):
    """Fusionne C6 (prioritaire pour la nature) et C7 (complément). {num: {...}}."""
    c6 = lire_c6(chemin_c6) if chemin_c6 else {}
    c7 = lire_c7(chemin_c7) if chemin_c7 else {}
    out = {num: dict(d) for num, d in c6.items()}
    for num, d in c7.items():
        if num not in out:
            out[num] = {"nature": d.get("nature", ""), "avant": d.get("avant", ""),
                        "apres": d.get("apres", ""), "nb": None, "lat": None, "lon": None,
                        "source": "C7"}
        elif not out[num].get("nature") and d.get("nature"):
            out[num]["nature"] = d["nature"]
    return out


# ---------------------------------------------------------------------------
# Jointure à la couche PT
# ---------------------------------------------------------------------------

def _est_poteau(row):
    s = str(row.get("TYPE_STRUC") or "").upper()
    return "POTEAU" in s or "POTELET" in s


def associer_pt(pt_gdf, annexes, tol_m=12.0):
    """Associe chaque poteau PT à un N° appui d'annexe.

    Renvoie {index_pt: num_appui}. Étape 1 : match par ``NOM``. Étape 2 : pour les
    appuis restants disposant de coordonnées, plus proche poteau PT dans ``tol_m``."""
    assoc = {}
    if pt_gdf is None or not annexes:
        return assoc

    poteaux = [(idx, row) for idx, row in pt_gdf.iterrows() if _est_poteau(row)]

    # 1) par numéro : NOM / CODE / dernier groupe de chiffres
    restants = dict(annexes)
    for idx, row in poteaux:
        for cand in _nums_candidats(row):
            if cand in annexes:
                assoc[idx] = cand
                restants.pop(cand, None)
                break

    # 2) par géolocalisation (appuis restants avec lat/lon)
    a_geo = [(num, d["lat"], d["lon"]) for num, d in restants.items()
             if d.get("lat") is not None and d.get("lon") is not None]
    if a_geo and pt_gdf.crs is not None:
        try:
            import geopandas as gpd
            from shapely.geometry import Point
            g = gpd.GeoDataFrame(
                {"num": [x[0] for x in a_geo]},
                geometry=[Point(x[2], x[1]) for x in a_geo], crs="EPSG:4326",
            ).to_crs(pt_gdf.crs)
            deja = set(assoc.keys())
            for _, gr in g.iterrows():
                best, bestd = None, tol_m
                for idx, row in poteaux:
                    if idx in deja or row.geometry is None:
                        continue
                    d = row.geometry.distance(gr.geometry)
                    if d < bestd:
                        best, bestd = idx, d
                if best is not None:
                    assoc[best] = gr["num"]
                    deja.add(best)
        except Exception as e:
            logger.warning(f"Association PT par géolocalisation impossible : {e}")
    return assoc


def _insee_depuis_pt(pt_gdf):
    """Devine le code INSEE depuis les CODE existants (FT_<insee>_<num>)."""
    if pt_gdf is None or "CODE" not in pt_gdf.columns:
        return ""
    for v in pt_gdf["CODE"].dropna().tolist():
        m = re.match(r"FT_(\d{4,6})_", str(v))
        if m:
            return m.group(1)
    return ""


def enrichir_pt_nom_code(pt_gdf, assoc, insee=None):
    """Met à jour UNIQUEMENT ``NOM`` + ``CODE`` des poteaux associés.

    Le schéma du SHP est préservé (aucun champ ajouté/retiré). Renvoie le nombre
    de poteaux mis à jour."""
    if pt_gdf is None or not assoc:
        return 0
    if insee is None:
        insee = _insee_depuis_pt(pt_gdf)
    n = 0
    for idx, num in assoc.items():
        change = False
        if "NOM" in pt_gdf.columns and _num(pt_gdf.at[idx, "NOM"]) != num:
            pt_gdf.at[idx, "NOM"] = num
            change = True
        if "CODE" in pt_gdf.columns:
            code_new = f"FT_{insee}_{num}" if insee else f"FT_{num}"
            if str(pt_gdf.at[idx, "CODE"]) != code_new:
                pt_gdf.at[idx, "CODE"] = code_new
                change = True
        if change:
            n += 1
    return n


def natures_par_nom(pt_gdf, annexes, assoc=None):
    """{NOM_poteau: nature} pour la Console / le PDF / la symbologie (par numéro)."""
    if pt_gdf is None or not annexes:
        return {}
    if assoc is None:
        assoc = associer_pt(pt_gdf, annexes)
    out = {}
    for idx, num in assoc.items():
        nat = (annexes.get(num) or {}).get("nature", "")
        if nat:
            nom = _num(pt_gdf.at[idx, "NOM"]) if "NOM" in pt_gdf.columns else num
            out[nom or num] = nat
    return out


def compter_natures(annexes, assoc):
    """Comptes {remplacer, recaler, renforcer} des appuis RÉELLEMENT associés au PT."""
    c = {"remplacer": 0, "recaler": 0, "renforcer": 0}
    for num in set(assoc.values()):
        nat = (annexes.get(num) or {}).get("nature", "")
        if nat == "REMPLACEMENT":
            c["remplacer"] += 1
        elif nat == "RECALAGE":
            c["recaler"] += 1
        elif nat == "RENFORCEMENT":
            c["renforcer"] += 1
    return c
